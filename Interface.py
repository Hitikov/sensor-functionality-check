﻿import csv
import os

import xlsxwriter
from openpyxl import load_workbook, styles

from Function_characteristics import *
from Graphics import *


def ReadFile(path):
    with open(path, newline='') as csvfile:
        filereader = csv.reader(csvfile, delimiter=',')
        return ReaderToList(filereader)
           
def ReaderToList(filereader):
    datalist = []
    for row in filereader:
        datalist.append(tuple(row))
    datalist = SliceData(datalist, 1500, 1600)
    datalist, blanks_detected = convert_to_float(datalist)
    return datalist

def XlsxOutput(datalist, filename):
    name = filename
    workbook = xlsxwriter.Workbook(name + '.xlsx')
    worksheet = workbook.add_worksheet()

    row = 0
    col = 0
    for item1, item2 in (datalist):
        worksheet.write(row, col, float(item1))
        worksheet.write(row, col + 1, float(item2))
        row += 1
    workbook.close()

def AnalyseData(datalist, xlsxpath, imgpath, record_name):
    #Intial create
    if not(os.path.exists(xlsxpath)):   
        workbook = xlsxwriter.Workbook(xlsxpath)
        worksheet = workbook.add_worksheet()

        #Field labels
        worksheet.write(0, 0, "Номер с примечанием")
        worksheet.write(0, 1, "Потери, дБ")
        worksheet.write(0, 2, "Контраст интерференции, дБ")
        worksheet.write(0, 3, "FSR")
        worksheet.write(0, 4, "Спектр")
        worksheet.write(0, 5, "Состояние")

        workbook.close()
    

    #Creating and saving the graphic
    img = CreateGraphic(datalist)
    img.save(imgpath)

    #Find first free row
    workbookLoader = load_workbook(filename = xlsxpath)
    sheet = workbookLoader.active
    freeRow = sheet.max_row + 1

    losses, contrast, FSR, tempvariable = dataset_analysis(datalist)
    spectre_link = imgpath
    if dataset_validation([losses, contrast], -20, 7):
        fillColor = styles.colors.Color(rgb='0099FF33')
        condition = "Годен"
    else:
        fillColor = styles.colors.Color(rgb='00FF2E2E')
        condition = "Негоден"
    data = [record_name, losses, contrast, FSR, spectre_link, condition]
    for i in range(6):
        cell = sheet.cell(freeRow, i + 1).value = data[i]
    fill = styles.fills.PatternFill(patternType='solid', fgColor=fillColor)
    for i in range(6):
        cell = sheet.cell(freeRow, i + 1).fill = fill

    workbookLoader.save(xlsxpath)

# Validation of records and results directories existence
def check_directory_requirements():
    relative_path_records = 'records'
    relative_path_results = 'results'
    if (not os.path.exists(relative_path_records)
        or not os.path.exists(relative_path_results)):
        return False

    return True

# Creation of records and results directories
def prepare_folders():
    relative_path_records = 'records'
    relative_path_results = 'results'
    os.makedirs(relative_path_records, exist_ok=True)
    os.makedirs(relative_path_results, exist_ok=True)

# Get list of file names in specified directory with specified extension
def get_file_names(path, extension):
    return [entry.name for entry in os.scandir(path) if entry.is_file() and entry.name.endswith(extension)]